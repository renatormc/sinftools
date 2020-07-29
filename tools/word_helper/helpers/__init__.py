from openpyxl import load_workbook, drawing
from openpyxl.utils.cell import get_column_letter
from pathlib import Path
import re
import shutil
# import xlwings
import win32com.client as win32
import pythoncom

def connect_excel():
    pythoncom.CoInitialize()
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True
    return excel


def analise_pics():
    objects = {}
    for entry in Path("fotos").iterdir():
        reg = re.compile(r'((^[A-Za-z]+)(\d+)).*')
        res = reg.search(entry.name)
        obj = res.group(1).upper()
        try:
            objects[obj]['pics'].append(str(entry))
        except KeyError:
            objects[obj] = {'number': int(res.group(3)), 'pics': [str(entry)]} 
    # objects.sort(key=lambda x: x['number'])
    items = []
    for key, value in objects.items():
        for pic in value['pics']:
            items.append({'name': key, 'number': value['number'], 'pic': pic})
    items.sort(key=lambda x: x['number'])
    return items


def get_nome_laudo(nome, tipo):
    reg = re.compile(r'(^[A-Za-z]+)(\d+)')
    res = reg.search(nome)
    number = int(res.group(2))
    return f"Evidência {number} - {tipo}"


def get_pessoas_envolvidas(objects):
    pessoas = []
    for name, obj in objects.items():
        if obj['owner'] and not obj['owner'] in pessoas:
            pessoas.append(obj['owner'])
    return pessoas
    

def write_pics_sheet():
    excel = connect_excel()
    if not excel:
        print("Não foi possível se conectar ao Excel")
        return
    wb = excel.ActiveWorkbook
    ws = wb.Worksheets("objetos")

    #Clear
    last_row = ws.Cells(ws.Rows.Count, "A").End(-4162).Row
    for i in range(2, last_row + 1):
        rgn = ws.Range(f"A{i}:Z{i}")
        rgn.Clear()
    for pic in ws.Pictures():
        pic.Delete()


    i = 2
    for item in analise_pics():
        ws.Rows(i).RowHeight = 100
        ws.Range(f"A{i}").Value = item['name']
        ws.Range(f"B{i}").Value = 'Outro'

        ws.Range(f"E{i}").Value = f"Evidência {item['number']}"
        path = Path(item['pic']).absolute()
        ws.Range(f"C{i}").Value = str(path)
        rng = ws.Range(f"D{i}")
        pic = ws.Pictures().Insert(str(path))
        pic.Left = rng.Left + 10
        pic.Top = rng.Top
        if pic.Width > pic.Height:
            pic.Width = 100
        else:
            pic.Height = 100
        i += 1
    wb.Save()


def read_objects():
    excel = connect_excel()
    if not excel:
        print("Não foi possível se conectar ao Excel")
        return
    wb = excel.ActiveWorkbook
    ws = wb.Worksheets("principal")
    ret = {}
    last_row = ws.Cells(ws.Rows.Count, "A").End(-4162).Row
    for i in range(1, last_row + 1):
        ret[ws.Range(f"A{i}").Value] = ws.Range(f"B{i}").Value
    ws = wb.Worksheets("objetos")
    objects = {}
    last_row = ws.Cells(ws.Rows.Count, "A").End(-4162).Row
    for i in range(2, last_row + 1):
        name = ws.Range(f"A{i}").Value
        try:
            objects[ws.Range(f"A{i}").Value]['pics'].append(
                {'path': ws.Range(f"C{i}").Value, 'caption': ws.Range(f"E{i}").Value})
        except KeyError:
            nome_laudo = get_nome_laudo(ws.Range(f"A{i}").Value, ws.Range(f"B{i}").Value)
            objects[ws.Range(f"A{i}").Value] = {
                'pics': [{'path': ws.Range(f"C{i}").Value, 'caption': ws.Range(f"E{i}").Value}],
                'owner': ws.Range(f"F{i}").Value,
                'lacre': ws.Range(f"G{i}").Value,
                'type': ws.Range(f"B{i}").Value,
                'nome_laudo': nome_laudo
            }
    pics = []
    for key, value in objects.items():
        for pic in value['pics']:
            pics.append(pic)
    ret['objects'] = objects
    ret['pics'] = pics
    return ret



def read_sheet():
    # wb = load_workbook(filename = 'data.xlsx')
    # ws = wb['principal']
    return read_objects()