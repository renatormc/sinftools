import config
import openpyxl as oxl
import re

class ColMap:
    name = 1
    type = 2
    pics = 3

class ExcelHandler:
    
    def scan_pics(self):
        wb = oxl.load_workbook(str(config.excel_data_file))
        ws = wb['Objetos']
        objs = self.get_objects_from_pics()
        for i, obj in enumerate(objs):
            row = i + 2
            ws.cell(row=row, column=ColMap.name).value = f"Evidência {obj['number']}"
            ws.cell(row=row, column=ColMap.type).value = "Celular"
            pics = ",".join(obj['pics'])
            print(pics)
            ws.cell(row=row, column=ColMap.pics).value = pics
            wb.save(str(config.excel_data_file))
        

    def get_objects_from_pics(self):
        objects = {}
        for entry in config.pics_folder.iterdir():
            reg = re.compile(r'((^[A-Za-z]+)(\d+)).*')
            res = reg.search(entry.name)
            obj = res.group(1).upper()
            try:
                objects[obj]['pics'].append(entry.name)
            except KeyError:
                objects[obj] = {'number': int(res.group(3)), 'pics': [entry.name]}
        items = []
        for key, value in objects.items():
            objects[key]['pics'].sort()
            items.append({'name': key, 'number': value['number'], 'pics': value['pics']})
        items.sort(key=lambda x: x['number'])
        return items


    def get_objects_info(self):
        wb = oxl.load_workbook(str(config.excel_data_file))
        ws = wb['Objetos']
        objs = []
        for row in ws.iter_rows():
            name = str(row[ColMap.name - 1].value).strip()
            type_ = str(row[ColMap.type - 1].value).strip()
            pics = str(row[ColMap.pics - 1].value).strip().split(",")
            pics = [pic.strip() for pic in pics]
            objs.append({'name': name, 'type': type_, 'pics': pics})
        return objs
