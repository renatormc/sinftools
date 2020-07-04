import os

from sinf.report.config_handler import ler_arquivo_configuracoes
from openpyxl import load_workbook


class Analisador:
    def __init__(self):
        self.erros = []
        self.config_custom = ler_arquivo_configuracoes()

    def checar_planilha(self, arquivo):
        print("Lendo arquivo excel")
        wb = load_workbook(filename=arquivo)
        lista_planilhas = wb.get_sheet_names()
        if self.config_custom['itens']['bate-papo']:
            print("Checando planilha de bate-papos")
            if not self.config_custom['planilhas']['bate_papos'] in lista_planilhas:
                self.erros.append(
                    f"Não foi encontrada nenhuma planilha de nome {self.config_custom['planilhas']['bate_papos']} no arquivo XLSX.")
            else:
                ws = wb[self.config_custom['planilhas']['bate_papos']]
                cabecalho = []
                for item in ws[int(self.config_custom['linha_cabecalho'])]:
                    cabecalho.append(item.value)
                #
                for coluna in self.config_custom['mapa_colunas'].values():
                    if not coluna in cabecalho:
                        self.erros.append(
                            f"Não foi encontrada a coluna de nome '{coluna}' na planilha '{self.config_custom['planilhas']['bate_papos']}' na linha {self.config_custom['linha_cabecalho']}.")

        if self.config_custom['itens']['sms']:
            print("Checando planilha de sms")
            if not self.config_custom['planilhas']['sms'] in lista_planilhas:
                self.erros.append(
                    f"Não foi encontrada nenhuma planilha de nome {self.config_custom['planilhas']['sms']} no arquivo XLSX.")
            else:
                ws = wb[self.config_custom['planilhas']['sms']]
                cabecalho = []
                for item in ws[int(self.config_custom['linha_cabecalho'])]:
                    cabecalho.append(item.value)
                #
                for coluna in self.config_custom['mapa_colunas_smss'].values():
                    if not coluna in cabecalho:
                        self.erros.append(
                            f"Não foi encontrada a coluna de nome '{coluna}' na planilha '{self.config_custom['planilhas']['sms']}' na linha {self.config_custom['linha_cabecalho']}.")

        if self.config_custom['itens']['contato']:
            print("Checando planilha de contatos")
            if not self.config_custom['planilhas']['contatos'] in lista_planilhas:
                self.erros.append(
                    f"Não foi encontrada nenhuma planilha de nome {self.config_custom['planilhas']['contatos']} no arquivo XLSX.")
            else:
                ws = wb[self.config_custom['planilhas']['contatos']]
                cabecalho = []
                for item in ws[int(self.config_custom['linha_cabecalho'])]:
                    cabecalho.append(item.value)
                #
                for coluna in self.config_custom['mapa_colunas_contatos'].values():
                    if not coluna in cabecalho:
                        self.erros.append(
                            f"Não foi encontrada a coluna de nome '{coluna}' na planilha '{self.config_custom['planilhas']['contatos']}' na linha {self.config_custom['linha_cabecalho']}.")

        if self.config_custom['itens']['chamada']:
            print("Checando planilha de chamadas")
            if not self.config_custom['planilhas']['chamadas'] in lista_planilhas:
                self.erros.append(
                    f"Não foi encontrada nenhuma planilha de nome {self.config_custom['planilhas']['chamadas']} no arquivo XLSX.")
            else:
                ws = wb[self.config_custom['planilhas']['chamadas']]
                cabecalho = []
                for item in ws[int(self.config_custom['linha_cabecalho'])]:
                    cabecalho.append(item.value)
                #
                for coluna in self.config_custom['mapa_colunas_chamadas'].values():
                    if not coluna in cabecalho:
                        self.erros.append(
                            f"Não foi encontrada a coluna de nome '{coluna}' na planilha '{self.config_custom['planilhas']['chamadas']}' na linha {self.config_custom['linha_cabecalho']}.")

        if self.config_custom['itens']['imagem']:
            print("Checando planilha de imagens")
            if not self.config_custom['planilhas']['imagens'] in lista_planilhas:
                self.erros.append(
                    f"Não foi encontrada nenhuma planilha de nome {self.config_custom['planilhas']['imagens']} no arquivo XLSX.")
            else:
                ws = wb[self.config_custom['planilhas']['imagens']]
                cabecalho = []
                for item in ws[int(self.config_custom['linha_cabecalho'])]:
                    cabecalho.append(item.value)
                #
                for coluna in self.config_custom['mapa_colunas_imagens'].values():
                    if not coluna in cabecalho:
                        self.erros.append(
                            f"Não foi encontrada a coluna de nome '{coluna}' na planilha '{self.config_custom['planilhas']['imagens']}' na linha {self.config_custom['linha_cabecalho']}.")

        if self.config_custom['itens']['video']:
            print("Checando planilha de vídeos")
            if not self.config_custom['planilhas']['videos'] in lista_planilhas:
                self.erros.append(
                    f"Não foi encontrada nenhuma planilha de nome {self.config_custom['planilhas']['videos']} no arquivo XLSX.")
            else:
                ws = wb[self.config_custom['planilhas']['videos']]
                cabecalho = []
                for item in ws[int(self.config_custom['linha_cabecalho'])]:
                    cabecalho.append(item.value)
                #
                for coluna in self.config_custom['mapa_colunas_videos'].values():
                    if not coluna in cabecalho:
                        self.erros.append(
                            f"Não foi encontrada a coluna de nome '{coluna}' na planilha '{self.config_custom['planilhas']['videos']}' na linha {self.config_custom['linha_cabecalho']}.")

if __name__ == "__main__":
    os.chdir(r'E:\dados\galaxy')
    analisador = Analisador()
    analisador.checar_planilha(r'E:\dados\galaxy\Relatório.xlsx')
    for erro in analisador.erros:
        print(erro)
