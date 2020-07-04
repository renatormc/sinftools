from openpyxl import load_workbook
import codecs
import json
import os

script_dir = os.path.dirname(os.path.realpath(__file__))

versao = "6.1.2.10"
arquivo = r'E:\dados\galaxy\Relatório.xlsx'

with codecs.open(f"{script_dir}\\config_files\\{versao}.json", "r", "utf-8") as f:
    data = json.load(f)
wb = load_workbook(arquivo)

colunas = {"bate-papos": [], "smss": [], "chamadas": [], "contatos": [], "imagens": [], "videos": []}

ws = wb[data['planilhas']['bate_papos']]
for item in ws[int(data['linha_cabecalho'])]:
    colunas['bate-papos'].append(item.value)

# ws = wb[data['planilhas']['sms']]
# for item in ws[int(data['linha_cabecalho'])]:
#     colunas['smss'].append(item.value)

ws = wb[data['planilhas']['chamadas']]
for item in ws[int(data['linha_cabecalho'])]:
    colunas['chamadas'].append(item.value)

ws = wb[data['planilhas']['contatos']]
for item in ws[int(data['linha_cabecalho'])]:
    colunas['contatos'].append(item.value)

ws = wb[data['planilhas']['imagens']]
for item in ws[int(data['linha_cabecalho'])]:
    colunas['imagens'].append(item.value)

ws = wb[data['planilhas']['videos']]
for item in ws[int(data['linha_cabecalho'])]:
    colunas['videos'].append(item.value)


with codecs.open(f"{script_dir}\\excel_columns\\{versao}.json", "w", "utf-8") as f:
    f.write(json.dumps(colunas, ensure_ascii=False, indent=2))
