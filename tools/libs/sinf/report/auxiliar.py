# -*- coding: utf-8 -*-
import errno
import os
import re
import shutil
import sqlite3

# from sinf.report.config import config


def copyanything(src, dst):
    """copia um diretório não vazio com todo o seu conteúdo para outro"""
    try:
        shutil.copytree(src, dst)
    except OSError as exc:  # python >2.5
        if exc.errno == errno.ENOTDIR:
            shutil.copy(src, dst)
        else:
            raise


def copyAllExistentDir(src, dest):
    for root, dirs, files in os.walk(src):
        folder = root.replace(src, "")
        criar = f"{dest}{folder}"
        if not os.path.exists(f"{dest}{folder}"):
            os.makedirs(criar, exist_ok=True)
        for file_ in files:
            shutil.copyfile(f"{root}\\{file_}", f"{dest}{folder}\\{file_}")


def formataIdWhatsapp(id):
    id = str(id).strip()
    res = re.search(r'((.*)@s.whatsapp.net)(.*)', id)
    if res:
        partes = []
        if res.groups()[1].strip() != '':
            partes.append(res.groups()[1].strip())
        if res.groups()[2].strip() != '':
            partes.append(res.groups()[2].strip())
        if len(partes) > 0:
            retorno = " - ".join(partes)
        else:
            retorno = id
    else:
        res = re.search(r'((.*)@g.us)(.*)', str(id))
        if res:
            if res.groups()[2].strip() != '':
                retorno = res.groups()[2].strip()
            else:
                retorno = f"Grupo - {res.groups()[1].strip()}"
                # retorno = id
        else:
            retorno = id
    retorno = retorno.strip()
    if retorno != '':
        return retorno
    else:
        return 'Desconhecido'


def getAvatar(identificador, tipotelefone):
    if tipotelefone == "android":
        return getAvatarAndroid(identificador)
    if tipotelefone == "iphone":
        return getAvatarIphone(identificador)


def getAvatarAndroid(identificador):
    res = re.search(r'((.*)@s.whatsapp.net)(.*)', identificador)
    if res:
        if res.groups()[0].strip() != '':
            return res.groups()[0].strip() + ".jpg"
        else:
            retorno = ''
    else:
        res = re.search(r'((.*)@g.us)(.*)', str(identificador))
        if res:
            if res.groups()[0].strip() != '':
                return res.groups()[0].strip() + ".jpg"
            else:
                retorno = ''
        else:
            return ''


def getAvatarIphone(identificador):
    if "@" in identificador:
        parts = identificador.split("@")
        return f"{parts[0]}.jpg"
    return f"{identificador}.jpg"


def getAvatarWhatsapp(row, campo, tipo_telefone):
    if tipo_telefone == 'android':
        if row['user_message'] and campo != 'chat_name':
            return 'Me.jpg'
        return getAvatarAndroid(row[campo])
    if tipo_telefone == 'iphone':
        # if row['user_message'] and campo != 'chat_name':
        #     return 'Photo.jpg'
        return getAvatarIphone(row[campo])


def formataIdFacebook(id):
    return str(id).strip().replace("ONE_TO_ONE:", "")


def formataId(id, aplicativo):
    if aplicativo == "WhatsApp":
        return formataIdWhatsapp(id)
    elif aplicativo == "Facebook messenger":
        return formataIdFacebook(id)
    else:
        return id


def loadFromExcel(arquivo, planilha, config_custom, linha_cabecalho=2, metodo=1, bate_papos=True, smss=True,
                  chamadas=True, contatos=True, imagens=False, videos=False, audios=False):
    """recebe o nome do arquivo excel a ser aberto e retorna um dataframe contendo os dados da planilha"""
    from openpyxl import load_workbook
    import pandas as pd
    wb = load_workbook(filename=arquivo)
    retorno = {
        "bate_papos": None,
        "smss": None,
        "contatos": None,
        "chamadas": None,
        "imagens": None,
        "videos": None,
        "audios": None
    }

    if bate_papos:
        ws = wb[planilha]
        # montar dicionario de colunas
        data = ws.values
        next(data)
        cabecalho = {}
        i = 0
        for item in next(data):
            cabecalho[item] = i + 1
            i += 1
        # ws.max_column
        # ws.cell(row=2, column=44).value

        # Montar dicionários de links
        links = {}
        for row in range(linha_cabecalho, ws.max_row + 1):
            valor = ws.cell(row=row, column=cabecalho[config_custom['mapa_colunas']['attachment']]).value
            if valor:
                try:
                    links[ws.cell(row=row, column=cabecalho[config_custom['mapa_colunas']['id']]).value] = ws.cell(row=row,
                                                                                                            column=
                                                                                                            cabecalho[
                                                                                                                config_custom[
                                                                                                                    'mapa_colunas'][
                                                                                                                    'attachment']]).hyperlink.target.replace(
                        "\\", "/")
                except:
                    links[ws.cell(row=row, column=cabecalho[
                        config_custom['mapa_colunas']['id']]).value] = "{} (anexo indisponível)".format(valor)

        if metodo == 1:
            df = pd.read_excel(arquivo, skiprows=linha_cabecalho - 1, encoding='utf-8')
            wb.close()
        elif metodo == 2:
            data = ws.values
            for i in range(int(config_custom['linha_cabecalho']) - 1):
                next(data)
            cols = next(data)
            data = list(data)
            df = pd.DataFrame(data, columns=cols)

        # Substitui nome do arquivo por endereco
        def modificarAnexo(row):
            """Pega o link em vez do texto existente na célula do excel"""
            try:
                retorno = links[row[config_custom['mapa_colunas']['id']]]
            except:
                retorno = ''
            return retorno

        df['attachment_link'] = df.apply(modificarAnexo, axis=1)
        retorno['bate_papos'] = df

    if smss:
        ws = wb[config_custom['planilhas']['sms']]
        data = ws.values
        for i in range(int(config_custom['linha_cabecalho']) - 1):
            next(data)
        cols = next(data)
        data = list(data)
        df = pd.DataFrame(data, columns=cols)
        retorno['smss'] = df

    if contatos:
        ws = wb[config_custom['planilhas']['contatos']]
        data = ws.values
        for i in range(int(config_custom['linha_cabecalho']) - 1):
            next(data)
        cols = next(data)
        data = list(data)
        df = pd.DataFrame(data, columns=cols)
        retorno['contatos'] = df

    if chamadas:
        ws = wb[config_custom['planilhas']['chamadas']]
        data = ws.values
        for i in range(int(config_custom['linha_cabecalho']) - 1):
            next(data)
        cols = next(data)
        data = list(data)
        df = pd.DataFrame(data, columns=cols)
        retorno['chamadas'] = df

    if imagens:
        ws = wb.get_sheet_by_name(config_custom['planilhas']['imagens'])

        cabecalho = {}
        for i, item in enumerate(ws[linha_cabecalho]):
            cabecalho[item.value] = i + 1
            i += 1

        # Montar dicionários de links
        links = {}
        for row in range(linha_cabecalho, ws.max_row + 1):
            valor = ws.cell(row=row, column=cabecalho[config_custom['mapa_colunas_imagens']['name']]).value
            if valor:
                try:
                    links[ws.cell(row=row, column=cabecalho[config_custom['mapa_colunas_imagens']['id']]).value] = \
                        ws.cell(row=row,  column=cabecalho[ config_custom['mapa_colunas_imagens']['name']]).hyperlink.target.replace("\\", "/")
                except:
                    links[ws.cell(row=row, column=cabecalho[
                        config_custom['mapa_colunas_imagens']['id']]).value] = "{} (anexo indisponível)".format(valor)

        if metodo == 1:
            df = pd.read_excel(arquivo, sheet_name=config_custom['planilhas']['imagens'], skiprows=linha_cabecalho - 1, encoding='utf-8')
            wb.close()
        elif metodo == 2:
            data = ws.values
            for i in range(int(config_custom['linha_cabecalho']) - 1):
                next(data)
            cols = next(data)
            data = list(data)
            df = pd.DataFrame(data, columns=cols)

         # Substitui nome do arquivo por endereco
        def modificarAnexo(row):
            """Pega o link em vez do texto existente na célula do excel"""
            try:
                retorno = links[row[config_custom['mapa_colunas_imagens']['id']]]
            except:
                retorno = ''
            return retorno

        df['link'] = df.apply(modificarAnexo, axis=1)
        retorno['imagens'] = df

    if audios:
        ws = wb.get_sheet_by_name(config_custom['planilhas']['audios'])

        cabecalho = {}
        for i, item in enumerate(ws[linha_cabecalho]):
            cabecalho[item.value] = i + 1
            i += 1

        # Montar dicionários de links
        links = {}
        for row in range(linha_cabecalho, ws.max_row + 1):
            valor = ws.cell(row=row, column=cabecalho[config_custom['mapa_colunas_audios']['name']]).value
            if valor:
                try:
                    links[ws.cell(row=row, column=cabecalho[config_custom['mapa_colunas_audios']['id']]).value] = \
                        ws.cell(row=row,  column=cabecalho[ config_custom['mapa_colunas_audios']['name']]).hyperlink.target.replace("\\", "/")
                except:
                    links[ws.cell(row=row, column=cabecalho[
                        config_custom['mapa_colunas_audios']['id']]).value] = "{} (anexo indisponível)".format(valor)

        if metodo == 1:
            df = pd.read_excel(arquivo, sheet_name=config_custom['planilhas']['audios'], skiprows=linha_cabecalho - 1, encoding='utf-8')
            wb.close()
        elif metodo == 2:
            data = ws.values
            for i in range(int(config_custom['linha_cabecalho']) - 1):
                next(data)
            cols = next(data)
            data = list(data)
            df = pd.DataFrame(data, columns=cols)


        # Substitui nome do arquivo por endereco
        def modificarAnexo(row):
            """Pega o link em vez do texto existente na célula do excel"""
            try:
                retorno = links[row[config_custom['mapa_colunas_audios']['id']]]
            except:
                retorno = ''
            return retorno

        df['link'] = df.apply(modificarAnexo, axis=1)
        retorno['audios'] = df

    if videos:
        ws = wb.get_sheet_by_name(config_custom['planilhas']['videos'])

        cabecalho = {}
        for i, item in enumerate(ws[linha_cabecalho]):
            cabecalho[item.value] = i + 1
            i += 1

        # Montar dicionários de links
        links = {}
        for row in range(linha_cabecalho, ws.max_row + 1):
            valor = ws.cell(row=row, column=cabecalho[config_custom['mapa_colunas_videos']['name']]).value
            if valor:
                
                try:
                    links[ws.cell(row=row, column=cabecalho[config_custom['mapa_colunas_videos']['id']]).value] = \
                        ws.cell(row=row,
                                column=cabecalho[config_custom['mapa_colunas_videos']['name']]).hyperlink.target.replace("\\",
                                                                                                                   "/")
                except:
                    links[ws.cell(row=row, column=cabecalho[
                        config_custom['mapa_colunas_videos']['id']]).value] = "{} (anexo indisponível)".format(valor)

        if metodo == 1:
            df = pd.read_excel(arquivo, sheet_name=config_custom['planilhas']['videos'], skiprows=linha_cabecalho - 1,
                               encoding='utf-8')
            wb.close()
        elif metodo == 2:
            data = ws.values
            for i in range(int(config_custom['linha_cabecalho']) - 1):
                next(data)
            cols = next(data)
            data = list(data)
            df = pd.DataFrame(data, columns=cols)

        # Substitui nome do arquivo por endereco
        def modificarAnexo(row):
            """Pega o link em vez do texto existente na célula do excel"""
    
            try:
                retorno = links[row[config_custom['mapa_colunas_videos']['id']]]
            except:
                retorno = ''
            return retorno
        
        df['link'] = df.apply(modificarAnexo, axis=1)
        retorno['videos'] = df

    return retorno


def eDono(de, dono_celular):
    if de:
        for conta in dono_celular.values():
            teste = conta.strip()
            if teste in str(de) and teste != '':
                return True
    return False


def getParticipantes(df_, tipo_telefone):
    """Recebe o dataframe filtrado de um grupo só e retorno uma lista contendo os participantes daquele grupo"""
    row = df_.iloc[len(df_) - 1]
    participantes = []
    parts = []
    if row['participants'] is not None and isinstance(row['participants'], str):
        parts = row['participants'].split('\n')
        for part in parts:
            part = part.strip()
            if part != "" and "@broadcast" not in part:
                participantes.append(
                    {'identificador_formatado': formataId(part, row['app']).strip(), 'avatar_participante': getAvatar(part, tipo_telefone)})
    return participantes


def getTodasMidias(df_, categoria):
    """Recebe o dataframe já filtrado por grupo e retorna a lista de todos os vídeos, audios, imagens, ou arquivos existentes"""
    df2_ = df_[df_['attachment_type'] == categoria]
    retorno = {'identificador_formatado': df_.iloc[0]['formatted_chat_name'], 'midias': []}
    if len(df2_) > 0:
        for idx, row in df2_.iterrows():
            retorno['midias'].append({'mensagem_id': row['id'], 'link': row['attachment_link']})
    return retorno


def get_last_avatar_names(dir_):
    files = os.listdir(dir_)
    contacts = []
    for filename in files:
        parts = filename.split("-")
        if len(parts) == 2:
            contact = parts[0]
        elif len(parts) == 3:
            contact = f"{parts[0]}-{parts[1]}"
        if contact not in contacts:
            contacts.append(contact)

    avatars = []
    for contact in contacts:
        avatar = None
        for item in filter(lambda x: x.startswith(contact), files):
            timestamp = item.replace(f"{contact}-","").replace(".jpg", "").replace(".thumb", "")
            try:
                timestamp = int(timestamp)
            except:
                continue
            if avatar is None:
                avatar = (item, timestamp, contact)
                continue
            if avatar[1] < timestamp:
                avatar = (item, timestamp, contact)
        avatars.append(avatar)
    return avatars


def converteIdentificadorFacebook(valor):
    reg = re.compile(r'\d+')
    result = reg.findall(valor)
    if result:
        return result[0]
    else:
        return valor


def formatarIdentificadorFacebook(id):
    if not os.path.exists('.report\\contacts_db2'):
        return id
    id_completo = id
    id = converteIdentificadorFacebook(id)
    try:
        cursor = sqlite3.connect('.report\\contacts_db2').cursor()
        cursor.execute("""
        SELECT display_name
        FROM contacts
        WHERE fbid=:Id
        """, {"Id": id})
        return cursor.fetchall()[0][0]
    except:
        try:
            cursor = sqlite3.connect('.report\\contacts_db2').cursor()
            cursor.execute("""
            SELECT display_name
            FROM contact_summaries
            WHERE fbid=:Id
            """, {"Id": id})
            return cursor.fetchall()[0][0]
        except:
            try:
                return id_completo
            except:
                return id


def getPage(df, total_pages=None, page=1, per_page=1500, only_count=False):
    count = df.shape[0]
    rest = df.shape[0] % per_page
    if not total_pages: #calcula somente se não foi informado
        if rest > 0:
            total_pages = int(df.shape[0] / per_page) + 1
        else:
            total_pages = int(df.shape[0] / per_page)
    if only_count:
        return total_pages
    first = (page - 1) * per_page
    last = first + per_page
    df2 = df.iloc[first:last]
    return {
        "count": count,
        "total_pages": total_pages,
        "per_page": per_page,
        "page": page,
        "df": df2,
        "first": first,
        "last": last
    }


if __name__ == "__main__":
    files = os.listdir(r'C:\Users\renato\Desktop\temp\Relatório Iphone\Apple_iPhone 7 Plus (A1784)\.report\Profile')
    res = get_last_avatar_names(files)
    for item in filter(lambda x: '2044' in x[2], res):
        print(item)