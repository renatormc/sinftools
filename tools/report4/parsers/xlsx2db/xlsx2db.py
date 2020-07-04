from openpyxl import load_workbook
from difflib import SequenceMatcher
import parsers.xlsx2db.settings as st
from models import *
from parsers.parser_base import ParserBase


class XlsxParser(ParserBase):
    def __init__(self, filename):
        self.filename = filename
        self.wb = None
        self.ws = None
        self.headers_row = None
        self.headers = None
        self.headers_mapped = None
        self.sheets = None
        self.selected_sheet = None
        self.found_sheets = None

    def load(self):
        self.wb = load_workbook(os.path.join(
            self.read_source.folder, self.filename))
        self.sheets = self.find_out_matches(
            st.possible_sheet_names, self.wb.sheetnames)
        self.found_sheets = [key for key,
                                     value in self.sheets.items() if value]

    def check_env(self):
        msgs = []
        if not self.filename:
            msgs.append("Nenhum arquivo XLSX encontrado")
        return msgs

    def run(self):
        self.load()
        self.parse_all()

    def get_sheets(self):
        return self.found_sheets

    def find_out_matches(self, possible_names, existing_names, limiar=0.7):
        ret = {}
        for key, candidates in possible_names.items():
            item = {}
            for existing_name in existing_names:
                greater_ratio = 0
                for candidate in candidates:
                    ratio = SequenceMatcher(None, existing_name.strip(
                    ).lower(), candidate.strip().lower()).ratio()
                    if ratio > greater_ratio:
                        greater_ratio = ratio
                item[existing_name] = greater_ratio
            probable = max(zip(item.values(), item.keys()))
            ret[key] = probable[1] if probable[0] >= limiar else None
        return ret

    def select_sheet(self, name):
        if name not in self.found_sheets:
            return
        self.ws = self.wb.get_sheet_by_name(self.sheets[name])

        # get headers row and headers names
        for i in range(10):
            n = 0
            for j in range(self.ws.max_column):
                if self.ws[i + 1][j].value:
                    n += 1
            if n > 0.7 * self.ws.max_column:
                self.headers_row = i + 1
                break
        if not self.headers_row:
            return

        self.headers = {column.value: i + 1 for i,
                                                column in enumerate(self.ws[self.headers_row])}
        possible_column_names = {key: value['possible_names']
                                 for key, value in st.columns[name].items()}
        fields = self.find_out_matches(possible_column_names, self.headers)

        self.headers_mapped = {key: self.headers[value]
                               for key, value in fields.items() if value}
        self.selected_sheet = name

    def read_row(self, row_number):
        row_number = row_number + self.headers_row
        row = {}
        for field in self.headers_mapped.keys():
            converter = st.columns[self.selected_sheet][field]['converter']
            row[field] = converter(self.ws.cell(
                row=row_number, column=self.headers_mapped[field]))
        return row

    def count_rows(self):
        return self.ws.max_row - self.headers_row

    def parse_contact(self):
        self.select_sheet('contact')
        for i in range(self.count_rows()):
            row = self.read_row(i + 1)
            contact = Contact()
            contact.name = row['name']
            contact.source = row['source']
            contact.deleted_state = row['deleted']
            for item in row['entries']:
                entry = ContactEntry()
                entry.category = item['category']
                entry.value = item['value']
                entry.read_id = self.read_id
                self.add(entry)
                contact.entries.append(entry)
            contact.read_id = self.read_id
            self.add(contact)
        self.commit()

    def parse_sms(self):
        self.select_sheet('sms')
        for i in range(self.count_rows()):
            row = self.read_row(i + 1)
            sms = Sms()
            sms.body = row['body']
            sms.folder = row['folder']
            sms.status = row['status']
            sms.timestamp = row['timestamp']
            for item in row['parties']:
                part = SmsPart()
                part.role = item['role']
                part.identifier = item['identifier']
                part.name = item['name']
                part.read_id = self.read_id
                self.add(part)
                sms.parties.append(part)
            sms.read_id = self.read_id
            self.add(sms)
        self.commit()

    def parse_call(self):
        self.select_sheet('call')
        for i in range(self.count_rows()):
            row = self.read_row(i + 1)
            call = Call()
            call.duration = row['duration']
            call.timestamp = row['timestamp']
            call.type_ = row['type']
            for item in row['parties']:
                part = CallPart()
                part.role = item['role']
                part.identifier = item['identifier']
                part.name = item['name']
                part.read_id = self.read_id
                self.add(part)
                call.parties.append(part)
            call.read_id = self.read_id
            self.add(call)
        self.commit()

    def parse_image(self):
        self.select_sheet('image')
        for i in range(self.count_rows()):
            row = self.read_row(i + 1)
            file = File()
            file.extracted_path = os.path.join(
                self.read_source.folder, row['extracted_path'])
            file.creation_time = row['creation_time']
            file.modify_time = row['modify_time']
            file.access_time = row['access_time']
            file.md5 = row['md5']
            file.sha256 = row['sha256']
            file.deleted_state = row['deleted']
            file.type_ = 'image'
            file.read_id = self.read_id
            self.add(file)
        self.commit()

    def parse_video(self):
        self.select_sheet('video')
        for i in range(self.count_rows()):
            row = self.read_row(i + 1)
            file = File()
            file.extracted_path = os.path.join(
                self.read_source.folder, row['extracted_path'])
            file.creation_time = row['creation_time']
            file.modify_time = row['modify_time']
            file.access_time = row['access_time']
            file.md5 = row['md5']
            file.sha256 = row['sha256']
            file.deleted_state = row['deleted']
            file.type_ = 'video'
            file.read_id = self.read_id
            self.add(file)
        self.commit()

    def parse_audio(self):
        self.select_sheet('audio')
        for i in range(self.count_rows()):
            row = self.read_row(i + 1)
            file = File()
            file.extracted_path = os.path.join(
                self.read_source.folder, row['extracted_path'])
            file.creation_time = row['creation_time']
            file.modify_time = row['modify_time']
            file.access_time = row['access_time']
            file.md5 = row['md5']
            file.sha256 = row['sha256']
            file.deleted_state = row['deleted']
            file.type_ = 'audio'
            self.add(file)
        self.commit()

    def add_participant(self, identifier, name):
        if not identifier and not name:
            return None
        participant = db_session.query(Participant).filter(
            Participant.identifier == identifier, Participant.name == name).first()
        if not participant:
            participant = Participant()
            participant.identifier = identifier
            participant.name = name
            participant.read_id = self.read_id
            self.add(participant)
            self.commit()
        return participant

    def parse_chat(self, min=0, max_=None):
        self.select_sheet('chat')
        chats_already_included = {}
        for i in range(self.count_rows()):
            if max_ and i > max_:
                break
            if i < min:
                continue
            row = self.read_row(i + 1)
            if row['chat_id'] not in chats_already_included.keys():
                chat = Chat()
                chat.deleted_state = row['deleted_chat']
                chat.start_time = row['start_time']
                chat.start_time = row['start_time']
                chat.identifier = row['chat_name']['identifier']
                chat.name = row['chat_name']['name']
                self.add(chat)
                chats_already_included[row['chat_id']] = chat
            for item in row['participants']:
                participant = self.add_participant(
                    item['identifier'], item['name'])
                if participant and participant not in chats_already_included[row['chat_id']].participants:
                    chats_already_included[row['chat_id']
                    ].participants.append(participant)
            message = Message()
            message.from_ = self.add_participant(
                row['from']['identifier'], row['from']['name'])
            message.body = row['body']
            message.deleted_state = row['deleted_message']
            message.timestamp = row['timestamp']

            if row['attachment']:
                attachment = File()
                attachment.extracted_path = os.path.join(
                    self.read_source.folder, row['attachment'])
                attachment.meta_data = row['attachment_details']
                message.attachments.append(attachment)
            message.read_id - self.read_id
            self.add(message)
            chats_already_included[row['chat_id']].messages.append(message)
        self.commit()

    def parse_user_account(self):
        self.select_sheet('user_account')
        for i in range(self.count_rows()):
            row = self.read_row(i + 1)
            user_account = UserAccount()
            user_account.name = row['name']
            user_account.username = row['username']
            user_account.password = row['password']
            user_account.service_type = row['service_type']
            user_account.deleted_state = row['deleted']
            user_account.read_id = self.read_id
            self.add(user_account)
        self.commit()

    def parse_all(self, min=0, max_=None):
        if 'sms' in self.found_sheets:
            self.parse_sms()
        if 'call' in self.found_sheets:
            self.parse_call()
        if 'contact' in self.found_sheets:
            self.parse_contact()
        if 'chat' in self.found_sheets:
            self.parse_chat(min=min, max_=max_)
        if 'image' in self.found_sheets:
            self.parse_image()
        if 'audio' in self.found_sheets:
            self.parse_audio()
        if 'video' in self.found_sheets:
            self.parse_video()
        if 'user_account' in self.found_sheets:
            self.parse_user_account()
