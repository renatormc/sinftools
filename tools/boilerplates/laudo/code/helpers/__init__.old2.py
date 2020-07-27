from openpyxl import load_workbook, drawing
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from pathlib import Path
import re
# import xlwings
import win32com.client


def connect_excel():
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True
    return excel


def analise_pics():
    objects = {}
    for entry in Path("fotos").iterdir():
        reg = re.compile(r'((^[A-Za-z]+)(\d+)).*')
        res = reg.search(entry.name)
        obj = res.group(1).upper()
        try:
            objects[obj]['pics'].append(str(entry))
        except KeyError:
            objects[obj] = {'number': int(res.group(3)), 'pics': [
                str(entry.absolute())]}
    # objects.sort(key=lambda x: x['number'])
    items = []
    for key, value in objects.items():
        for pic in value['pics']:
            items.append({'name': key, 'number': value['number'], 'pic': pic})
    items.sort(key=lambda x: x['number'])
    return items


def write_pics_sheet():
    wb = load_workbook(filename='helpers\\data.xlsx')
    ws = wb['objetos']
    i = 2
    for item in analise_pics():
        ws[f"A{i}"].value = item['name']
        ws[f"B{i}"].value = "Outro"
        ws[f"C{i}"].value = item['pic']
        img = drawing.image.Image(item['pic'])
        img.anchor = f"D{i}"
        ws.row_dimensions[i].height = 100
        img.height = 100
        img.width = 100
        ws.add_image(img)

        ws[f"E{i}"].value = f"Evidência {item['number']}"
        i += 1
    wb.save(filename='data.xlsx')


def read_sheet():
    wb = load_workbook(filename='data.xlsx')
    ws = wb['principal']
    ret = {}
    for row in ws.iter_rows():
        ret[row[0].value] = row[1].value
    objects = {}
    ws = wb['objetos']
    for i in range(2, ws.max_row + 1):
        try:
            path = Path()
            objects[ws[f"A{i}"].value]['pics'].append(
                {'path': ws[f"C{i}"].value, 'caption': ws[f"E{i}"].value})
        except KeyError:
            objects[ws[f"A{i}"].value] = {
                'pics': [{'path': ws[f"C{i}"].value, 'caption': ws[f"E{i}"].value}],
                'owner': ws[f"F{i}"].value,
                'lacre': ws[f"G{i}"].value,
                'type': ws[f"B{i}"].value
            }
        ret['objects'] = objects
    return ret
